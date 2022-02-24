# -*- coding: utf-8 -*-
"""
Created on Wed Jul  3 22:46:57 2019

@author: Sleepymouse
"""

import glob
import numpy as np
import pandas as pd
import openpyxl
import pylab
from scipy.optimize import curve_fit



jet_lag = 20
a = 11 #初めの11行
b = jet_lag * 24 * 60 #recoveryの行数
gain = [] #計算したgainの入れ物
wb = openpyxl.Workbook()

#好きな場所に好きな名前のフォルダを作り、その中にdouble_plot.pyと解析したい.xlsファイル1つをおく
path = glob.glob("./*.xls") #xlsファイルを見つける。
df = pd.read_excel(path[0], sheet_name = 0, header = None)
Ch = np.array(df.iloc[10, 1:])


df = df.drop(0, axis=1)
df = df.drop(range(0,11,1), axis=0)
df = df.reset_index(drop=True)


num = len(df.columns)
time_points = len(df.index)
recdays = int(len(df) / (24 * 60))


#マウス1匹ごとに処理していく
for ID in range(num): #エクセル上で一番左のマウスから処理していく    

    column = (np.array(df.iloc[:, ID])).astype(np.float64)
    c = np.argwhere(np.isnan(column))
    for d in np.reshape(c, (-1)):
        if d <= time_points-10:
            column[d] = np.floor(np.nanmean(column[(d - 10):(d + 11)]))
        elif d > time_points-10:
            column[d] = np.floor(np.nanmean(column[(d - 10):]))
#    df.iloc[:,i] = column
        
        
    onset = [] #これが周期長データを格納していく
    
    xdata = list(range(-10,10))#x軸の準備     
    wb.create_sheet(title = str(Ch[ID])) #シートの追加

    
    #一日分のarrayの取得
    for day in range(jet_lag): #jetlag1日目を0としている
        deltastock = [] #差を格納するリストの準備
        
        #ZT3-30の区間内で、ある点を起点とした前後6時間の増加量を計算
        n = 60 * 3 - 1 #起点
        m = 60 * 3 - 1 #起点
        
        if day == 0:
            oneday = column[(day * 60 * 24):(day * 60 * 24 + 60 * 27)] #初日のみ30時間分のデータを取得。
            
            for i in range(60 * 21): #試行回数の設定
                if n <= (60 * 24): #起点がZT24までについて、以下の計算をする
                    list1 = oneday[n - (60 * 3 - 1):n ] #起点マイナス3時間のデータを含むリスト
                    list2 = oneday[n:n + (60 * 3 - 1)] #起点プラス3時間のデータを含むリスト
                    sum1 = sum(list1) #起点マイナス3時間の和
                    sum2 = sum(list2) #起点プラス3時間の和
                    delta = sum2 - sum1 #増加量を計算
                    deltastock.append(delta) #計算した増加量をdeltastockリストにためておく
                    n = n + 1
                    
            maxindex = np.argmax(deltastock) + 60*3 #増加量が最大であるindexを取得。起点の時間のズレ(60*3)を補正
            onset.append(maxindex)
                  

        else:
            oneday = column[(day * 60 * 24 - 60 * 3):(day * 60 * 24 + 60 * 27)] #36時間分のデータを取得。
            
            for i in range(60 * 24): #試行回数の設定
                if n <= (60 * 24): #起点がZT24までについて、以下の計算をする
                    list1 = oneday[m - (60 * 3 - 1):m ] #起点マイナス3時間のデータを含むリスト
                    list2 = oneday[m:m + (60 * 3 - 1)] #起点プラス3時間のデータを含むリスト
                    sum1 = sum(list1) #起点マイナス3時間の和
                    sum2 = sum(list2) #起点プラス3時間の和
                    delta = sum2 - sum1 #増加量を計算
                    deltastock.append(delta) #計算した増加量をdeltastockリストにためておく
                    m = m + 1 #1分ずつずらしてもう一度計算へ
            
            maxindex = np.argmax(deltastock)#増加量が最大であるindexを取得
            onset.append(maxindex)
            
    maxi = np.max(onset)
    onset_re = onset/ maxi
    
    
    ws = wb.get_sheet_by_name(str(Ch[ID])) #番号に対応したシートを選択する

    #エクセルファイルへデータの書き込み
    ws['A1'].value = 'activity onset'
    
    for p in range(len(onset)):
        ws.cell(row = 2, column = p + 1).value = xdata[p]
        ws.cell(row = 3, column = p + 1).value = onset[p]        
    
    
    def sigmoid(x, x0, k, a, c):
        y = a / (1 + np.exp(-k*(x-x0))) + c
        return y
 
    sigmoid_format = r'$y = \frac{%f}{1 + e^{-%f (x - %f)}} + %f$'
 
    popt, pcov = curve_fit(sigmoid, xdata, onset_re)
    print("Result : x0=%f, k=%f, a=%f, c=%f" % (popt[0], popt[1], popt[2], popt[3]))
    
    x = np.linspace(-10, 10) 
    yopt = sigmoid(x, *popt)
    
    pylab.plot(xdata, onset_re, 'o')
    pylab.plot(x, yopt, label=('Optimized : ' + (sigmoid_format % (popt[2], popt[1], popt[0], popt[3]))) )
    pylab.ylim(0, 1)
    pylab.xlim(-10, 10)
    pylab.legend(loc='lower left')
    pylab.grid(True)
    name = str(Ch[ID]) + ".png"
    pylab.savefig(name)
    pylab.show()
        
    gain.append(popt[1])
    
    ws['A9'].value = "gain" 
    ws['A10'].value = popt[1] #回帰直線の傾きをエクセルに記入

    

#gainを書き込むシートの準備
ws = wb.get_sheet_by_name("Sheet") #一番手前のシートを選択
ws.title = "gain" #シート名をgainに変更

#データの書き込み
for q in range(len(gain)):
     ws.cell(row = 1, column = q + 1).value = Ch[q]
     ws.cell(row = 2, column = q + 1).value = gain[q]

wb.save("phase_shift.xls") #エクセルの保存

print(str(num) + "匹のマウスに対して処理が終わりました-")
    

